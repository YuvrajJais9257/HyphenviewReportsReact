import re
from datetime import datetime
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A1
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.drawing.image import Image as excel_img
from io import BytesIO
import mysql.connector
import psycopg2
 

def html_text_convertor(html):
    try:
 
        text = re.sub(r'<br\s*/?>', ' ', html)
        text = re.sub(r'<style[^>]*>.*?</style>|<script[^>]*>.*?</script>', '', text)
        text = re.sub(r'<.*?>', '', text)
        text = re.sub(r'&\w+;', ' ', text)
        return text
 
    except Exception as err:
        raise ValueError(f"Error: {err}")
 
def correct_rows(row):
    modified_row = []
    for item in row:
        # print(type(item))
        if isinstance(item, str) and ('<' in item or '>' in item):
            new_item = html_text_convertor(item)
            modified_row.append(new_item)
        elif item == None:
            modified_row.append('None')
        elif isinstance(item, list):
            modified_row.append(str(item))
        elif isinstance(item, datetime):
            modified_row.append(item.replace(tzinfo=None))
        else:
            modified_row.append(item)
    return modified_row

def split_text(text, max_length):
    return [text[i:i + max_length] for i in range(0, len(text), max_length)]

def detect_long_text_columns(rows, max_length=10):
    long_text_columns = []
    for key in rows[0].keys():
        for row in rows:
            if isinstance(row[key], str) and len(row[key]) > max_length:
                long_text_columns.append(key)
                break
    return long_text_columns
 
def detect_long_text_columns_postgres(rows, columns, max_length=10):
    """
    Detect long text columns in PostgreSQL result with explicit column information.
    """
    long_text_columns = []
    keys = [col.name for col in columns]
    for i, key in enumerate(keys):
        for row in rows:
            if isinstance(row[i], str) and len(row[i]) > max_length:
                long_text_columns.append(key)
                break
 
    return long_text_columns

def calculate_column_widths_mysql(page_width, num_columns, rows, min_absolute_width=80):
    """
    Calculate column widths dynamically based on page width, number of columns, and content.
    """
    flag = 0
    max_column_widths = [len(col) for col in rows[0].keys()]
    for row in rows:
        for i, key in enumerate(row.keys()):
            max_width = len(str(row[key]))
            max_column_widths[i] = max(max_width, max_column_widths[i])
 
    total_width = sum(max_column_widths)
 
    while total_width < page_width:
        scaling_factor = (page_width) / total_width
        max_column_widths = [max(min_absolute_width, (width * scaling_factor)) for width in max_column_widths]
        total_width = sum(max_column_widths)
        flag = 1
 
    if flag == 1:
        column_widths = [max(min_absolute_width, width) for width in max_column_widths]
    else:
        if total_width > page_width:
            scaling_factor = 0.6 * (page_width / total_width)
            column_widths = [max(min_absolute_width, width * scaling_factor) for width in max_column_widths]
    # Ensure that each column has a minimum width of min_absolute_width
 
    # Adjust font size based on the number of columns
    # max_font_size = 12
    # min_font_size = 8
    # font_size = max(min_font_size, min(max_font_size, 0.5 * (page_width / (num_columns * max_font_size))))
 
    return column_widths
 
def calculate_column_widths_postgres(page_width, num_columns, rows, min_absolute_width=80):
    """
    Calculate column widths dynamically based on page width, number of columns, and content for PostgreSQL cursor result.
    """
    flag = 0
    max_column_widths = [len(str(value)) for value in rows[0]]
    for row in rows:
        for i, value in enumerate(row):
            max_width = len(str(value))
            max_column_widths[i] = max(max_width, max_column_widths[i])
    total_width = sum(max_column_widths)
    count = 0
    while total_width < page_width:
        count +=1
        scaling_factor = (page_width) / total_width
        max_column_widths = [max(min_absolute_width, (width * scaling_factor)) for width in max_column_widths]
        total_width = sum(max_column_widths)
        flag = 1
 
    if flag == 1:
        column_widths = [max(min_absolute_width, width) for width in max_column_widths]
    else:
        if total_width > page_width:
            scaling_factor = 0.6 * (page_width / total_width)
            column_widths = [max(min_absolute_width, int(width * scaling_factor)) for width in max_column_widths]
 
    return column_widths

def first_page_content(canvas, doc, report_title,dates):
    dates = dates.strftime("%Y-%m-%d %H:%M:%S")
    page_width, page_height = doc.pagesize
    x_text = 350
    y_text = A1[0] - 40
    logo_path = 'logo.png'
    canvas.setFont("Helvetica-Bold", 15)
    canvas.drawString(x_text, y_text, report_title)
   
    x_text = A1[1] - 200
    canvas.setFont("Helvetica-Bold", 15)
    canvas.drawString(x_text, y_text, dates)
    canvas.drawImage(logo_path, 80, page_height - 100, width=200, height=80)

def generate_pdf_report(report_query, filename, db_config, title_text,db_type):
    try:
        print('@@@@@@@@@@@@@@@@@@@@@@@',db_config,'------------------------------')
        time = datetime.now()
        new_formatted_time = time.strftime("%Y-%m-%d_%H-%M-%S")
        # db_config = {
        #     'host': host,
        #     'user': username,
        #     'password': password,
        #     'database': database,
        #     'port' : port
        # }
 
        if db_type == 'mysql':
            conn = mysql.connector.connect(**db_config)
 
            cursor = conn.cursor(dictionary=True)
            cursor.execute(report_query)
 
            batch_size = 1000
            page_width, page_height = A1
            long_text_columns = None
            col_widths = None
            all_tables = []
            column_count = len(cursor.description)
            if column_count <= 12:
                for batch_count, result_batch in enumerate(iter(lambda: cursor.fetchmany(batch_size), [])):
                    if not result_batch:
                        break
 
                    if long_text_columns is None:
                        long_text_columns = detect_long_text_columns(result_batch)
                        col_widths = calculate_column_widths_mysql(page_width, len(result_batch[0]), result_batch)
 
                    table_data = []
                    keys = list(result_batch[0].keys())
                    table_data.append(tuple(keys))
 
                    font_size = 8
                    for col in long_text_columns:
                        col_index = keys.index(col)
                        max_content_length = 3000 # Adjust the max_content_length for cell
                        line_length = int(0.19 * col_widths[col_index])
                        for row in result_batch:
                            row[col] = '\n'.join(split_text(str(row[col])[:max_content_length], line_length))  # Adjust the max_length as needed
 
                    for row in result_batch:
                        new_row = list(row.values())
                        modified_row = correct_rows(new_row)
                        table_data.append(tuple(modified_row))
 
                    table = Table(table_data, colWidths=col_widths)
                    style = TableStyle([
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), font_size),
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ])
 
                    table.setStyle(style)
 
                    all_tables.append(table)
            else:
                return{"status":500,"Error":"Column limit exceeded, max columns allowed = 12"}
 
        elif db_type == 'postgres':
            conn = psycopg2.connect(**db_config)
 
            cursor = conn.cursor()
            cursor.execute(report_query)
 
            batch_size = 1000
            page_width, page_height = A1
            long_text_columns = None
            col_widths = None
            all_tables = []
            column_count = len(cursor.description)
            if column_count <= 12:
                for batch_count, result_batch in enumerate(iter(lambda: cursor.fetchmany(batch_size), [])):
                    if not result_batch:
                        break
                    long_text_columns = detect_long_text_columns_postgres(result_batch, cursor.description)
                    col_widths = calculate_column_widths_postgres(page_width, len(result_batch[0]), result_batch)
                    table_data = []
                    font_size = 8
 
                    keys = [desc[0] for desc in cursor.description]
                    table_data.append(keys)
 
                    for col in long_text_columns:
                        col_index = keys.index(col)
                        max_content_length = 3000
                        line_length = int(0.19 * col_widths[col_index])
                        result_batch = [list(row) for row in result_batch]
 
                        for row in result_batch:
                            row[col_index] = '\n'.join(split_text(str(row[col_index])[:max_content_length], line_length))
 
                    for row in result_batch:
                        modified_row = correct_rows(row)
                        table_data.append(tuple(modified_row))
 
                    table = Table(table_data, colWidths=col_widths)
                    style = TableStyle([
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), font_size),
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ])
 
                    table.setStyle(style)
 
                    all_tables.append(table)
            else:
                return{"status":500,"Error":"Column limit exceeded, max columns allowed = 12"}
        # valid_filename = f"{filename}_{new_formatted_time}"
        valid_filename = f"{filename}"
        def on_first_page_function(canvas, doc):
            first_page_content(canvas, doc, title_text, time)
 
        output_doc = SimpleDocTemplate(valid_filename, pagesize=(A1[1], A1[0]))
        output_doc.build(all_tables, onFirstPage=on_first_page_function)
        cursor.close()
        conn.close()
        # time = datetime.now()
        # new_formatted_time = time.strftime("%Y-%m-%d_%H-%M-%S")
        # # Set up PDF
        # doc = SimpleDocTemplate(f"{filename}", pagesize=(A1[1], A1[0]))
 
        # # Title
        # title = Paragraph(title_text, ParagraphStyle(name='Title', fontSize=16, alignment=1))
 
        # logo = Image('logo.png',width=100,height=100)
 
        # # From and To dates
        # from_date = datetime.now().strftime("%Y-%m-%d")
        # to_date = datetime.now().strftime("%Y-%m-%d")
        # dates = Paragraph(f"From: {from_date} To: {to_date}", ParagraphStyle(name='Date', alignment=1))
 
        # conn = mysql.connector.connect(**db_config)
 
        # cursor = conn.cursor(dictionary=True)
        # cursor.execute(report_query)
 
        # batch_size = 1000
        # page_width, page_height = A1
        # long_text_columns = None
        # col_widths = None
        # all_tables = []
 
        # for batch_count, result_batch in enumerate(iter(lambda: cursor.fetchmany(batch_size), [])):
        #     if not result_batch:
        #         break
 
        #     if long_text_columns is None:
        #         long_text_columns = detect_long_text_columns(result_batch)
        #         col_widths = calculate_column_widths_mysql(page_width, len(result_batch[0]), result_batch)
 
        #     table_data = []
        #     keys = list(result_batch[0].keys())
        #     table_data.append(tuple(keys))
 
        #     font_size = 8
        #     for col in long_text_columns:
        #         col_index = keys.index(col)
        #         max_content_length = 3000 # Adjust the max_content_length for cell
        #         line_length = int(0.19 * col_widths[col_index])
        #         for row in result_batch:
        #             row[col] = '\n'.join(split_text(str(row[col])[:max_content_length], line_length))  # Adjust the max_length as needed
 
        #     for row in result_batch:
        #         new_row = list(row.values())
        #         table_data.append(tuple(new_row))
 
        #     table = Table(table_data, colWidths=col_widths)
        #     style = TableStyle([
        #         ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        #         ('FONTSIZE', (0, 0), (-1, -1), font_size),
        #         ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        #         ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        #         ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        #         ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        #         ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        #         ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        #         ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        #         ('GRID', (0, 0), (-1, -1), 1, colors.black),
        #     ])
 
        #     table.setStyle(style)
 
        #     all_tables.append(table)
        # def on_first_page_function(canvas, doc,dates):
        #     first_page_content(canvas, doc, title_text,dates)
        # doc.build(all_tables, onFirstPage=on_first_page_function)
        # # doc.build(elements)
        # cursor.close()
        # conn.close()
        # print("PDF saved successfully.")

    except Exception as e:
        print(f"Error: {str(e)}")
 
def generate_excel_report(df,filename, title_text):
    df.columns = [col.replace('_', ' ').title() for col in df.columns]
 
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
 
    # Add logo
    logo_path = "logo.png"
    img = excel_img(logo_path)
    img.width = 100
    img.height = 50
    ws.add_image(img, 'A1')  # Add image to cell A1
    ws.row_dimensions[1].height = 50  # Adjust row height to accommodate the logo
 
    # Add title
    title_col_range = f"B1:{chr(64 + len(df.columns))}1"  # Determine column range for title
    ws.merge_cells(title_col_range)  # Merge cells for title
    title_cell = ws['B1']  # Assuming the title starts from column B
    title_cell.value = title_text
    # title_cell.alignment = Alignment(horizontal='center', vertical='center')  # Center align the title
    title_cell.font = Font(bold=True,size=16)
 
    # Add from & to date
    from_date = "From: 2024-01-01"
    to_date = "To: 2024-03-20"
    ws.merge_cells('B2:C2')
    ws['B2'] = from_date
    ws.merge_cells('D2:E2')
    ws['D2'] = to_date
    ws['B2'].alignment = Alignment(horizontal='center')
    ws['D2'].alignment = Alignment(horizontal='center')
 
    # Write DataFrame to Excel
    for r_idx, row in enumerate(df.iterrows(), start=4):
        for c_idx, value in enumerate(row[1], start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = Alignment(horizontal='center')
 
    # Apply borders to data
    border_style = Border(left=Side(border_style='thin'),
                          right=Side(border_style='thin'),
                          top=Side(border_style='thin'),
                          bottom=Side(border_style='thin'))
 
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border_style
 
    # Set column names (headers)
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = border_style
 
    # Set column widths
    for col_idx, column in enumerate(ws.columns, start=1):
        max_length = max(len(str(cell.value)) for cell in column)
        adjusted_width = (max_length + 2) * 1.2
        column_letter = chr(64 + col_idx)  # Convert column index to letter
        ws.column_dimensions[column_letter].width = adjusted_width
 
    # Save the workbook
    wb.save(filename)